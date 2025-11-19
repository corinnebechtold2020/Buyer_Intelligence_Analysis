import io
from datetime import datetime
import re

import numpy as np
import pandas as pd
import streamlit as st

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------------
# LSC Buyer Journey Intelligence Engine
# ------------------------

# Internal canonical required fields (after normalization)
REQUIRED_COLUMNS = [
    "reader_company",
    "first_name",
    "last_name",
    "activity_date",
    "activity_source",
    "content_title",
    "content_bj_label",
    "specifics",
    "content_topics",
    "content_source",
]

STAGE_ORDER = ["Awareness", "Problem Definition", "Solution Exploration", "Vendor Evaluation"]

STAGE_COLOR = {
    "Vendor Evaluation": "FF0000",  # red
    "Solution Exploration": "FFA500",  # orange
    "Problem Definition": "FFFF00",  # yellow
    "Awareness": "C0C0C0",  # gray
}


### --- Helpers for v5-style logic ---


def extract_primary_topic(row):
    """Return a comma-separated primary topic string based on content_topics, content_title, specifics."""
    parts = []
    for c in ["content_topics", "content_title", "specifics"]:
        val = row.get(c) if isinstance(row, dict) else (row.get(c) if c in row else None)
        if pd.isna(val) or val is None or str(val).strip() == "":
            continue
        toks = re.split(r"[,;|/\n]", str(val))
        toks = [t.strip() for t in toks if t and len(t.strip()) > 1]
        parts.extend(toks)

    if not parts:
        return ""
    freq = {}
    for p in parts:
        freq[p] = freq.get(p, 0) + 1
    top = sorted(freq.items(), key=lambda x: (-x[1], x[0]))[:3]
    return ",".join([t[0] for t in top])


VENDOR_LOOKUP = {
    "manufacturing": ["Cytiva", "Sartorius", "Thermo Fisher Scientific", "Repligen", "Pall (Cytiva)", "MilliporeSigma"],
    "qa/qc": ["Waters", "Agilent", "Thermo Fisher Scientific", "Shimadzu", "SCIEX", "Bruker"],
    "upstream": ["Cytiva", "Sartorius", "Repligen", "Eppendorf", "PBS Biotech"],
    "downstream": ["Cytiva", "Sartorius", "Tosoh Bioscience", "Repligen"],
    "analytical": ["Waters", "Agilent", "Thermo Fisher Scientific", "SCIEX", "Shimadzu"],
    "all molecules": ["Cytiva", "Sartorius", "Thermo Fisher Scientific", "MilliporeSigma"],
}


def map_topics_to_vendors(primary_topic_str: str):
    if not primary_topic_str:
        return ["Cytiva", "Sartorius", "Thermo Fisher Scientific", "MilliporeSigma", "Repligen", "Pall (Cytiva)"]
    vendors = []
    toks = [t.strip().lower() for t in primary_topic_str.split(",") if t.strip()]
    for t in toks:
        for k, vs in VENDOR_LOOKUP.items():
            if k in t or t in k:
                vendors.extend(vs)
    if not vendors:
        return ["Cytiva", "Sartorius", "Thermo Fisher Scientific", "MilliporeSigma", "Repligen"]
    seen = set()
    out = []
    for v in vendors:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


def extract_vendors_evaluated(row):
    text = " ".join([str(row.get(c, "")) for c in ["content_source", "specifics", "client", "newsletter_name"] if not pd.isna(row.get(c, ""))])
    if not text:
        return ""
    found = set()
    possible = [
        "Cytiva", "Sartorius", "Thermo Fisher", "Thermo Fisher Scientific", "MilliporeSigma", "Pall", "Waters", "Agilent",
        "SCIEX", "Eppendorf", "Fujifilm Irvine Scientific", "Repligen", "Lonza", "Catalent", "Bio-Rad", "PBS Biotech",
        "Tosoh", "Shimadzu",
    ]
    for p in possible:
        if re.search(re.escape(p), text, flags=re.IGNORECASE):
            found.add(p)
    return ", ".join(sorted(found))


def detect_behavioral_triggers_and_flags(group_df: pd.DataFrame):
    rows_text = " ".join([str(x) for x in group_df.get("content_title", []) if not pd.isna(x)]) + " " + " ".join([str(x) for x in group_df.get("specifics", []) if not pd.isna(x)])
    flags = []
    triggers = set()
    explanation_parts = []

    # Determine non-newsletter rows based ONLY on the `activity_source` text.
    # Requirement: any flags or explanation text derived from activity/source
    # should appear only when the activity source itself is a non-empty, non-newsletter value.
    # Treat empty or missing activity_source as unknown and DO NOT count it as non-newsletter.
    asrc = group_df.get("activity_source", pd.Series([""] * len(group_df))).fillna("").astype(str).str.lower().str.strip()
    # strict match for the word 'newsletter' to avoid partial matches
    is_news = asrc.str.contains(r"\bnewsletter\b", na=False)
    # Only treat rows as non-newsletter WHEN the activity_source explicitly contains
    # an explicit non-NL marker. This follows the user's instruction that generic
    # sources like 'webinar', 'event', or 'engagement' should NOT be considered Non-NL.
    non_nl_marker = asrc.str.contains(r"\b(?:non[\s_-]?nl|non[\s-]?newsletter|nonnewsletter)\b", na=False)
    non_nl = group_df[non_nl_marker]

    # tuned heuristic (conservative): mark Non-NL only when it's a strong signal.
    total_rows = len(group_df)
    non_nl_count = len(non_nl)
    pct_non_nl = (non_nl_count / total_rows) if total_rows > 0 else 0.0

    # compute process and late-stage signals on the full set of activity text (all rows).
    # Rationale: webinars/events and other generic activity_source values should still
    # surface process/late-stage interest signals even when they are NOT explicitly
    # labeled as Non-NL. However, the explicit "Non-NL engagement" wording should
    # only be added when explicit non-nl markers exist in `activity_source`.
    combined_all_text = " ".join([str(x) for x in group_df.get("content_title", []) if not pd.isna(x)]) + " " + " ".join([str(x) for x in group_df.get("specifics", []) if not pd.isna(x)])
    proc_hits = len(re.findall(r"(?:upstream|downstream|process|scale-up|tech transfer|manufactur|chromatograph|viral clearance|polish|isolation|purification)", combined_all_text, flags=re.IGNORECASE))
    late_hits = 1 if re.search(r"(?:compare|comparison|vs\b|pricing|quote|rfp|vendor evaluation|vendor eval|vendor|evaluation|spec sheet|datasheet|benchmark|compare to|comparison study)", combined_all_text, flags=re.IGNORECASE) else 0

    # If any explicit non-newsletter marker appears in `activity_source`, mark Non-NL.
    # This follows the user's requirement: only report Non-NL when the uploaded
    # file explicitly contains non-newsletter markers in the activity source.
    bj_vals = [str(x).lower() for x in group_df.get("content_bj_label", []) if not pd.isna(x)]
    combined_text = " ".join([str(x) for x in group_df.get("content_title", []) if not pd.isna(x)]) + " " + " ".join([str(x) for x in group_df.get("specifics", []) if not pd.isna(x)])
    inferred_stage = None
    if any("vendor" in v or "evaluation" in v or "eval" in v for v in bj_vals):
        inferred_stage = "Vendor Evaluation"
    elif re.search(r"(compare|comparison|vs\b|pricing|quote|rfp|vendor|evaluation)", combined_text, flags=re.IGNORECASE):
        inferred_stage = "Vendor Evaluation"
    elif re.search(r"(solution|product|datasheet|whitepaper|case study|webinar|feature|use case|workflow)", combined_text, flags=re.IGNORECASE):
        inferred_stage = "Solution Exploration"
    elif re.search(r"(problem|pain|challenge|need|gap|issue|requirement|diagnos)", combined_text, flags=re.IGNORECASE):
        inferred_stage = "Problem Definition"
    else:
        inferred_stage = "Awareness"

    # tuned thresholds
    # Strict behavior: if any row's `activity_source` explicitly matches the
    # non-newsletter pattern, add the Non-NL flag and explanation. Do not infer
    # Non-NL indirectly from other activity types.
    if non_nl_count >= 1:
        flags.append("\U0001F525 Non-NL engagement")
        explanation_parts.append("Direct (Non-NL) visit rather than newsletter click")

    late_stage_keywords = r"(?:compare|comparison|vs\b|pricing|quote|rfp|vendor evaluation|vendor eval|vendor|evaluation|spec sheet|datasheet|benchmark|compare to|comparison study)"
    # evaluate late-stage and process keywords on the full activity text so that
    # webinars/events and other generic activity sources can still trigger process
    # and late-stage flags. Note: the explicit Non-NL wording is still gated by
    # the non_nl markers above (cond_high/cond_stage) and will not be set unless
    # explicit non-newsletter rows exist.
    search_text_for_stage = combined_all_text
    if re.search(late_stage_keywords, search_text_for_stage, flags=re.IGNORECASE):
        flags.append("\U0001F4C4 Late-stage content")
        triggers.add("Vendor/option comparison")
        explanation_parts.append("Engaging with late-stage comparison/evaluation content")

    process_keywords = r"(?:upstream|downstream|process|scale-up|tech transfer|manufactur|chromatograph|viral clearance|polish|isolation|purification)"
    if len(re.findall(process_keywords, search_text_for_stage, flags=re.IGNORECASE)) >= 2:
        flags.append("\U0001F504 Deep process interest")
        triggers.add("GMP / clinical manufacturing")
        explanation_parts.append("Exploring process/solution workflows")

    problem_keywords = r"(?:problem|pain|challenge|issue|need|gap|failure|root cause)"
    # only consider problem/pain keywords in non-newsletter activity text
    if search_text_for_stage and re.search(problem_keywords, search_text_for_stage, flags=re.IGNORECASE):
        explanation_parts.append("Looking at challenge/problem-focused content")

    expl = []
    for part in explanation_parts:
        if part not in expl:
            expl.append(part)

    return {
        "flags": ", ".join(flags),
        "behavioral_triggers": ", ".join(sorted(triggers)),
        "explanation_parts": expl,
    }


def enrich_external_trigger_for_company(company_name: str):
    if not company_name or pd.isna(company_name):
        return ""
    cn = str(company_name).lower()
    if "samsung" in cn:
        return "Announced a large-scale CDMO expansion program in APAC."
    if "biocon" in cn:
        return "Inaugurated its first U.S. manufacturing facility, expanding global biologics production footprint."
    if "biontech" in cn:
        return "Investing in mRNA manufacturing capacity, including new mRNA facility initiatives."
    return ""


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize incoming LSC exports to a fixed internal schema.

    - Accepts multiple header variants and coalesces into canonical internal names
    - Errors only if required internal fields (REQUIRED_COLUMNS) are missing after normalization
    - Preserves extra columns
    - Creates optional columns when absent
    - Converts `activity_date` to datetime and strips whitespace
    """

    canonical_map = {
        # Reader Company
        "reader_company": "reader_company",
        "reader company": "reader_company",
        "reader_org": "reader_company",
        "reader org": "reader_company",
        "company": "reader_company",
        "company_name": "reader_company",

        # Names
        "first_name": "first_name",
        "firstname": "first_name",
        "last_name": "last_name",
        "lastname": "last_name",

        # Title
        "title": "title",

        # Dates
        "activity_date": "activity_date",
        "activity date": "activity_date",
        "activity_month": "activity_date",
        "engagement_date": "activity_date",

        # Activity / Source
        "activity_source": "activity_source",
        "activity source": "activity_source",
        "activity_type": "activity_source",
        "engagement_type": "activity_source",

        # Content
        "content_title": "content_title",
        "content title": "content_title",
        "raw_content_title": "content_title",
        "content_type": "content_type",
        "content type": "content_type",
        "content_buyer's_journey": "content_bj_label",
        "content buyer's journey": "content_bj_label",
        "content_bj_label": "content_bj_label",
        "content_topics": "content_topics",
        "content topics": "content_topics",
        "topics": "content_topics",
        "content_source": "content_source",
        "content source": "content_source",
        "specifics": "specifics",

        # Optional fields
        "email": "email",
        "newsletter_name": "newsletter_name",
        "newsletter name": "newsletter_name",
        "client": "client",
        "quality": "quality",
        "country": "country",
        "region": "region",
    }

    # group detected source columns to their canonical target
    norm_to_sources = {}
    for orig_col in list(df.columns):
        norm = orig_col.strip().lower().replace(" ", "_")
        target = canonical_map.get(norm)
        if target:
            norm_to_sources.setdefault(target, []).append(orig_col)

    # coalesce or rename
    for target, sources in norm_to_sources.items():
        if len(sources) == 1:
            df = df.rename(columns={sources[0]: target})
        else:
            df[target] = df[sources].bfill(axis=1).iloc[:, 0]

    # verify required internal fields
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        examples = [
            "Reader Company", "First Name", "Last Name", "Title", "Activity Date",
            "Activity Source", "Content Title", "Content Buyer's Journey", "Specifics",
            "Content Topics", "Content Source",
        ]
        raise ValueError(f"Uploaded file is missing required fields after normalization: {missing}. Examples: {examples}")

    # optional fallbacks
    optional_cols = ["email", "newsletter_name", "quality", "product", "client", "country", "region"]
    for c in optional_cols:
        if c not in df.columns:
            df[c] = np.nan

    # convert dates
    if "activity_date" in df.columns:
        df["activity_date"] = pd.to_datetime(df["activity_date"], errors="coerce")

    # strip whitespace
    obj_cols = df.select_dtypes(include=[object]).columns.tolist()
    for c in obj_cols:
        df[c] = df[c].apply(lambda v: v.strip() if isinstance(v, str) else v)

    return df


def find_required_columns(df):
    df = normalize_columns(df)
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    return missing, df


def parse_dates(df, col="engagement_date"):
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def stage_from_text(text: str):
    if pd.isna(text):
        return None
    t = str(text).lower()
    # heuristics
    if re.search(r"(demo|trial|pricing|rfp|quote|proposal|purchase|evaluation|vendor|vendor eval)", t):
        return "Vendor Evaluation"
    if re.search(r"(solution|product|datasheet|whitepaper|case study|webinar|feature|use case)", t):
        return "Solution Exploration"
    if re.search(r"(problem|pain|challenge|need|gap|issue|requirement|define|diagnos)", t):
        return "Problem Definition"
    if re.search(r"(newsletter|social|blog|press|awareness|announcement|advertis|marketing)", t):
        return "Awareness"
    # default fallback: Awareness
    return "Awareness"


@st.cache_data
def process_dataframe(df: pd.DataFrame, min_engagements: int = 5):
    """Process normalized DataFrame into v5-style Individuals, Company_Combined and Sales_Hot_List.

    This function recreates the transformation, explanation, primary topic extraction,
    vendor mapping, vendor-evaluated extraction, behavioral triggers, and company-level
    summaries similar to the provided v5 reference.
    """

    # ensure expected internals exist
    for c in ["email", "first_name", "last_name", "reader_company", "activity_source", "product"]:
        if c not in df.columns:
            df[c] = np.nan

    # canonical person key: prefer email, else first+last+company
    df["email_norm"] = df["email"].astype(str).str.lower().str.strip()

    def person_key(row):
        em = str(row.get("email_norm", "")).strip()
        if em and em.lower() not in ["nan", "none", ""]:
            return em
        return (str(row.get("first_name", "")).strip() + "|" + str(row.get("last_name", "")).strip() + "|" + str(row.get("reader_company", "")).strip()).lower()

    df["person_key"] = df.apply(person_key, axis=1)

    # compute per-person aggregates
    grouped = df.groupby("person_key")

    records = []
    for key, g in grouped:
        g = g.sort_values(by=["activity_date"]) if "activity_date" in g.columns else g
        first = g.iloc[0]
        total_eng = int(g["activity_date"].count()) if "activity_date" in g.columns else g.shape[0]

        # Primary topic from group
        primary_topic = extract_primary_topic(g.iloc[0] if not g.empty else {})
        if not primary_topic:
            # try aggregate across rows
            primary_topic = extract_primary_topic({"content_topics": ", ".join([str(x) for x in g.get("content_topics", []) if not pd.isna(x)]),
                                                   "content_title": ", ".join([str(x) for x in g.get("content_title", []) if not pd.isna(x)]),
                                                   "specifics": ", ".join([str(x) for x in g.get("specifics", []) if not pd.isna(x)])})

        potential_vendors = map_topics_to_vendors(primary_topic)
        vendors_eval = extract_vendors_evaluated({
            "content_source": ", ".join([str(x) for x in g.get("content_source", []) if not pd.isna(x)]),
            "specifics": ", ".join([str(x) for x in g.get("specifics", []) if not pd.isna(x)]),
            "client": ", ".join([str(x) for x in g.get("client", []) if not pd.isna(x)]),
            "newsletter_name": ", ".join([str(x) for x in g.get("newsletter_name", []) if not pd.isna(x)]),
        })

        bt = detect_behavioral_triggers_and_flags(g)

        # determine buyer journey label using multi-factor rules
        combined_text = " ".join([str(x) for x in g.get("content_title", []) if not pd.isna(x)]) + " " + " ".join([str(x) for x in g.get("specifics", []) if not pd.isna(x)])
        label = "Awareness"
        # prefer explicit content_bj_label values if present in group
        bj_vals = [str(x).lower() for x in g.get("content_bj_label", []) if not pd.isna(x)]
        if any("vendor" in v or "evaluation" in v or "eval" in v for v in bj_vals):
            label = "Vendor Evaluation"
        elif re.search(r"(compare|comparison|vs\b|pricing|quote|rfp|vendor|evaluation)", combined_text, flags=re.IGNORECASE):
            label = "Vendor Evaluation"
        elif re.search(r"(solution|product|datasheet|whitepaper|case study|webinar|feature|use case|workflow)", combined_text, flags=re.IGNORECASE):
            label = "Solution Exploration"
        elif re.search(r"(problem|pain|challenge|need|gap|issue|requirement|diagnos)", combined_text, flags=re.IGNORECASE):
            label = "Problem Definition"
        else:
            label = "Awareness"

        # explanation assembly
        expl_parts = []
        if bt.get("explanation_parts"):
            expl_parts.extend(bt.get("explanation_parts"))
        if label == "Awareness" and not expl_parts:
            expl_parts.append("General interest in this topic area")

        explanation = " \u2022 ".join(expl_parts) if expl_parts else "General interest in this topic area"

        reason = ""
        if label == "Vendor Evaluation":
            reason = f"Actively comparing solutions in {primary_topic}; good timing for targeted outreach."
        else:
            reason = f"Early awareness in {primary_topic}; keep warm with high-level thought leadership."

        records.append({
            "Reader Company": first.get("reader_company", ""),
            "First Name": first.get("first_name", ""),
            "Last Name": first.get("last_name", ""),
            "Title": first.get("title", ""),
            "Total Engagements": total_eng,
            "Buyer_Journey_Label": label,
            "Explanation": explanation,
            "Primary_Topic": primary_topic,
            "Potential_Vendors": ", ".join(potential_vendors),
            "Vendors_Evaluated": vendors_eval,
            "Buyer_Intensity_Flags": bt.get("flags", ""),
            "Behavioral_Trigger": bt.get("behavioral_triggers", ""),
            "External_Trigger (press release info, etc)": enrich_external_trigger_for_company(first.get("reader_company", "")),
            "Reason to Reach out": reason,
        })

    individuals_df = pd.DataFrame.from_records(records)

    # ensure numeric
    if "Total Engagements" in individuals_df.columns:
        individuals_df["Total Engagements"] = individuals_df["Total Engagements"].astype(int)

    # filter by min engagements
    individuals_filtered = individuals_df[individuals_df["Total Engagements"] >= int(min_engagements)].copy()

    # Company combined
    if individuals_filtered.empty:
        company_combined = pd.DataFrame(columns=["Company", "Active_Individuals", "Dominant_Topic", "Team_Buying_Signal", "Behavioral_Triggers", "External_Trigger", "Product_Category"])
    else:
        comp_grp = individuals_filtered.groupby("Reader Company")
        comp_rows = []
        for comp, cg in comp_grp:
            active_inds = cg.shape[0]
            # dominant topic: most common Primary_Topic among individuals
            dom = cg["Primary_Topic"].value_counts().idxmax() if cg["Primary_Topic"].notna().any() else ""
            team_signal = "Yes" if active_inds >= 2 else "No"
            beh = ", ".join(sorted(set(cg["Behavioral_Trigger"].dropna().astype(str).tolist())))
            ext = enrich_external_trigger_for_company(comp)
            # product category: union of primary topics across company
            pcs = []
            for p in cg["Primary_Topic"].dropna().astype(str):
                for tok in p.split(","):
                    t = tok.strip()
                    if t and t not in pcs:
                        pcs.append(t)
            prod_cat = ",".join(pcs)
            comp_rows.append({
                "Company": comp,
                "Active_Individuals": active_inds,
                "Dominant_Topic": dom,
                "Team_Buying_Signal": team_signal,
                "Behavioral_Triggers": beh,
                "External_Trigger": ext,
                "Product_Category": prod_cat,
            })
        company_combined = pd.DataFrame.from_records(comp_rows)

    # Sales hot list: individuals in Vendor Evaluation
    sales_hot = individuals_filtered[individuals_filtered["Buyer_Journey_Label"] == "Vendor Evaluation"].copy()
    if not sales_hot.empty:
        # adapt column names to expected Sales_Hot_List names
        sales_hot = sales_hot.rename(columns={
            "Total Engagements": "Total_Engagements",
            "Buyer_Journey_Label": "Buyer_Journey_Label",
            "Reason to Reach out": "Reason_to_Reach_Out",
        })
        # ensure column order matches v5 preview
        cols = [
            "Reader Company",
            "First Name",
            "Last Name",
            "Title",
            "Total_Engagements",
            "Buyer_Journey_Label",
            "Explanation",
            "Primary_Topic",
            "Potential_Vendors",
            "Vendors_Evaluated",
            "Buyer_Intensity_Flags",
            "Behavioral_Trigger",
            "External_Trigger (press release info, etc)",
            "Reason_to_Reach_Out",
        ]
        for c in cols:
            if c not in sales_hot.columns:
                sales_hot[c] = ""
        sales_hot = sales_hot[cols]

    # product_map: reuse earlier approach but simpler
    product_map = pd.DataFrame(columns=["product", "mentions", "unique_companies"]) if df.empty else df.groupby("product").agg(mentions=("product", "count"), unique_companies=("reader_company", lambda s: s.dropna().nunique())).reset_index().sort_values(by="mentions", ascending=False)

    # company_summary mirrors company_combined plus additional columns
    company_summary = company_combined.copy()

    return {
        "individuals": individuals_filtered,
        "company_combined": company_combined,
        "company_summary": company_summary,
        "product_map": product_map,
        "sales_hot_list": sales_hot,
        "raw": df,
    }


def excel_with_styles(dfs: dict, filename: str = "report.xlsx") -> bytes:
    # Writes multiple dataframes to an in-memory workbook and applies color-coding
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet, df in dfs.items():
            # sanitize sheet name
            sheet_name = sheet[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    output.seek(0)
    wb = load_workbook(output)

    # apply coloring rules
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # find header -> find buyer_stage or stage or total_engagements
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_index = {h: i + 1 for i, h in enumerate(headers) if h}

        # color rows by buyer_stage or Buyer_Journey_Label column if present
        if "buyer_stage" in header_index or "Buyer_Journey_Label" in header_index:
            col_key = "buyer_stage" if "buyer_stage" in header_index else "Buyer_Journey_Label"
            col_idx = header_index[col_key]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                stage = cell.value
                if stage in STAGE_COLOR:
                    fill = PatternFill(start_color=STAGE_COLOR[stage], end_color=STAGE_COLOR[stage], fill_type="solid")
                    cell.fill = fill

        # highlight high engagement rows (>=12) if any total engagements column present
        te_idx = None
        for key in ("total_engagements", "Total Engagements", "Total_Engagements"):
            if key in header_index:
                te_idx = header_index[key]
                break
        if te_idx:
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=te_idx).value
                try:
                    if val is not None and int(val) >= 12:
                        # apply green to the whole row
                        green = PatternFill(start_color="00C853", end_color="00C853", fill_type="solid")
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=c).fill = green
                except Exception:
                    pass

    # write back to bytes
    out2 = io.BytesIO()
    wb.save(out2)
    return out2.getvalue()


def enrich_companies_stub(companies: pd.DataFrame) -> pd.DataFrame:
    # Placeholder: integrate external trigger APIs here. Adds a column with a sample trigger flag.
    companies = companies.copy()
    companies["external_triggers"] = "(stub - no data)"
    return companies


def main():
    st.set_page_config(page_title="LSC Buyer Journey Intelligence Engine", layout="wide")
    st.title("LSC Buyer Journey Intelligence Engine")

    st.markdown(
        "Upload an XLSX or CSV export from Life Science Connect. The app validates columns, aggregates engagements to individuals, computes buyer journey stages, and produces an Excel report."
    )

    uploaded = st.file_uploader("Upload Life Science Connect XLSX/CSV", type=["xlsx", "csv"], accept_multiple_files=False)

    min_eng = st.sidebar.number_input("Minimum engagements per individual (filter)", value=5, min_value=1, step=1)
    enrich = st.sidebar.checkbox("Enrich high-intent companies (stub)", value=False)

    if uploaded is None:
        st.info("Upload a .xlsx or .csv file to begin. You can also open the sample below.")
        st.download_button("Download sample CSV", data=sample_csv(), file_name="sample_lsc_export.csv")
        return

    # Read file
    try:
        if uploaded.type == "text/csv" or uploaded.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded)
        else:
            df = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Failed to read file: {e}")
        return

    # Normalize columns immediately after upload and show normalized column list for debugging
    try:
        df = normalize_columns(df)
    except ValueError as e:
        st.error(str(e))
        st.write("Columns found:", list(df.columns))
        return

    st.write("Normalized Columns:", df.columns.tolist())

    # Process
    with st.spinner("Processing data — this may take a moment for tens of thousands of rows..."):
        results = process_dataframe(df, min_engagements=min_eng)

    st.success("Processing complete")

    # Show previews
    st.subheader("Individuals (filtered)")
    st.dataframe(results["individuals"].head(200))

    st.subheader("Company Summary")
    st.dataframe(results["company_summary"].head(200))

    st.subheader("Top Products")
    st.dataframe(results["product_map"].head(200))

    st.subheader("Sales Hot List")
    st.dataframe(results["sales_hot_list"].head(200))

    # optional enrichment
    if enrich:
        with st.spinner("Running enrichment (stub)..."):
            enriched = enrich_companies_stub(results["company_summary"])  # placeholder
            st.write(enriched.head(50))
            results["company_summary"] = enriched

    # Prepare Excel
    sheets = {
        "Company_Combined": results["company_combined"],
        "Company_Summary": results["company_summary"],
        "Individuals": results["individuals"],
        "Product_Map": results["product_map"],
        "Sales_Hot_List": results["sales_hot_list"],
    }

    excel_bytes = excel_with_styles(sheets, filename="LSC_Buyer_Journey_Report.xlsx")

    st.download_button(label="Download Excel Report", data=excel_bytes, file_name="LSC_Buyer_Journey_Report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def sample_csv() -> bytes:
    sample = pd.DataFrame(
        {
            "First Name": ["Alice", "Bob"],
            "Last Name": ["A", "B"],
            "Email": ["alice@example.com", "bob@example.com"],
            "Reader Company": ["Acme Inc", "Beta LLC"],
            "Activity Date": [datetime.now().isoformat(), datetime.now().isoformat()],
            "Activity Source": ["webinar - product features", "requested demo"],
            "Content Title": ["Webinar: Product X Features", "Requested Demo for Product Y"],
            "Product": ["Product X", "Product Y"],
        }
    )
    return sample.to_csv(index=False).encode("utf-8")


if __name__ == "__main__":
    main()
